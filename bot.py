import html
import random
import telebot
import re
import time
import os
from openpyxl import load_workbook
from datetime import datetime
from telebot.types import InlineQueryResultArticle, InputTextMessageContent
from collections import defaultdict
import json
import threading
from zoneinfo import ZoneInfo


TOKEN = (os.getenv("BOT_TOKEN") or "").strip()
if not TOKEN:
    raise RuntimeError("BOT_TOKEN is not set in environment variables")
BOT_USERNAME = "PhilosophyOracleBot"  # без @, например: PhilosophyOracleBot
EXCEL_FILE = "quotes.xlsx"

bot = telebot.TeleBot(TOKEN)
bags = defaultdict(list)
BOT_TZ = ZoneInfo("Europe/Moscow")
CHAT_STATE_FILE = "chat_state.json"

# сюда можно вставить file_id смешных стикеров
FUNNY_STICKER_SETS = [
    "cwcwhab_by_fStikBo",
    "monke2004",
    "JackalCats",
    "set_3099_by_makestick3_bot",
    "Yellowboi",
    "dog_1_by_MoiStikiBot",
    "ketrinsky",
    "HypeFoodByUffchat",
]

sticker_cache = {}

def esc(s: str) -> str:
    return html.escape(str(s), quote=False)

def load_chat_state():
    if not os.path.exists(CHAT_STATE_FILE):
        return {"known_chats": [], "daily_sent": {}}

    try:
        with open(CHAT_STATE_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            return {
                "known_chats": data.get("known_chats", []),
                "daily_sent": data.get("daily_sent", {})
            }
    except Exception:
        return {"known_chats": [], "daily_sent": {}}


def save_chat_state():
    data = {
        "known_chats": list(known_chats),
        "daily_sent": daily_sent
    }
    with open(CHAT_STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


chat_state = load_chat_state()
known_chats = set(chat_state["known_chats"])
daily_sent = chat_state["daily_sent"]


def today_str():
    return datetime.now(BOT_TZ).strftime("%Y-%m-%d")


def current_hm():
    now = datetime.now(BOT_TZ)
    return now.hour, now.minute


def remember_chat(chat_id: int):
    if chat_id not in known_chats:
        known_chats.add(chat_id)
        save_chat_state()


def was_daily_sent_today(chat_id: int) -> bool:
    return daily_sent.get(str(chat_id)) == today_str()


def mark_daily_sent(chat_id: int):
    daily_sent[str(chat_id)] = today_str()
    save_chat_state()

def get_random_sticker_file_id():
    if not FUNNY_STICKER_SETS:
        return None

    set_name = random.choice(FUNNY_STICKER_SETS)

    try:
        if set_name in sticker_cache:
            stickers = sticker_cache[set_name]
        else:
            sticker_set = bot.get_sticker_set(set_name)
            stickers = sticker_set.stickers
            sticker_cache[set_name] = stickers

        if not stickers:
            return None

        sticker = random.choice(stickers)
        return sticker.file_id

    except Exception as e:
        print(f"Failed to load sticker set {set_name}: {e}")
        return None

def pick_quote_bag(key: int):
    bag = bags[key]
    if not bag:
        bag = QUOTES[:]          # копия
        random.shuffle(bag)      # перемешали
        bags[key] = bag
    return bag.pop()

def load_quotes_from_excel(path: str):
    wb = load_workbook(path)
    ws = wb.active  # первый лист

    quotes = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # предполагаем, что 1-я строка = заголовки
        quote, author, source, tag = row[:4]

        # пропускаем пустые строки
        if not quote or not author:
            continue

        quotes.append({
            "quote": str(quote).strip(),
            "author": str(author).strip(),
            "source": str(source).strip() if source else "",
            "tag": str(tag).strip() if tag else ""
        })

    return quotes


QUOTES = load_quotes_from_excel(EXCEL_FILE)



def normalize_tag(raw: str) -> str:
    """ 'жизнь как последняя' -> 'жизнь_как_последняя' """
    if not raw:
        return ""
    t = str(raw).strip().lower()
    t = t.lstrip("#").strip()
    # заменяем запятые/точки/слэши на пробел, потом пробелы на _
    t = re.sub(r"[.,/]+", " ", t)
    t = re.sub(r"\s+", "_", t)
    # оставляем буквы/цифры/подчёркивания
    t = re.sub(r"[^0-9a-zа-яё_]+", "", t)
    t = re.sub(r"_+", "_", t).strip("_")
    return t

def strip_outer_quotes(text: str) -> str:
    """Убирает внешние кавычки, если они уже есть, чтобы не было ««... »»."""
    if not text:
        return ""
    t = str(text).strip()

    quote_pairs = [
        ("«", "»"),
        ('"', '"'),
        ("“", "”"),
        ("„", "“"),
        ("'", "'"),
        ("‚", "‘"),
        ("‹", "›"),
        ("(", ")"),
    ]

    changed = True
    while changed and len(t) >= 2:
        changed = False
        for lq, rq in quote_pairs:
            if t.startswith(lq) and t.endswith(rq):
                t = t[1:-1].strip()
                changed = True
                break

    return t



def emoji_for_tag(tag: str) -> str:
    for key, emoji in TAG_EMOJI.items():
        if key in tag:
            return emoji
    return "✨"

TAG_EMOJI = {
    # Спокойствие и внутренний покой
    "спокой": "🫧",
    "покой": "🫧",
    "тишин": "🫧",
    "невовлеч": "🫧",

    # Самоконтроль и фокус
    "самоконтрол": "🎯",
    "дисциплин": "🎯",
    "границ": "🎯",
    "фокус": "🎯",
    "вниман": "🎯",

    # Разум и ясность
    "разум": "🧠",
    "ясност": "🧠",
    "мышлен": "🧠",
    "объектив": "🧠",
    "перспектив": "🧠",
    "суждени": "🧠",

    # Мудрость и истина
    "мудрост": "🦉",
    "истин": "🦉",
    "философ": "🦉",

    # Принятие и судьба
    "принят": "🍃",
    "судьб": "🍃",
    "неизбеж": "🍃",
    "перемен": "🍃",

    # Стойкость и внутренняя опора
    "стойк": "🪨",
    "устойчив": "🪨",
    "опора": "🪨",
    "сила": "🪨",

    # Мужество и достоинство
    "мужеств": "🛡️",
    "смелост": "🛡️",
    "достоинств": "🛡️",

    # Добродетель и справедливость
    "справедлив": "⚖️",
    "добродетел": "⚖️",
    "нравствен": "⚖️",

    # Честность и искренность
    "честност": "🪞",
    "искрен": "🪞",
    "совест": "🪞",

    # Сострадание и доброта
    "сострадан": "🤍",
    "эмпат": "🤍",
    "добр": "🤍",

    # Общность и служение
    "общност": "🤝",
    "единств": "🤝",
    "служен": "🤝",
    "общее": "🤝",

    # Свобода и независимость
    "свобод": "🕊️",
    "независ": "🕊️",
    "самодостат": "🕊️",

    # Время и настоящее
    "врем": "⏳",
    "настоящ": "⏳",
    "мгновен": "⏳",

    # Смертность и бренность
    "смерт": "⚰️",
    "бренн": "⚰️",

    # Развитие и обучение
    "развит": "📚",
    "обучен": "📚",
    "саморазвит": "📚",

    # Действие и долг
    "действ": "🛠️",
    "долг": "🛠️",
    "труд": "🛠️",

    # Гнев и тревога
    "гнев": "🌋",
    "тревог": "🌋",
    "обид": "🌋",

    # Благодарность
    "благодар": "🙏",

    # Цель и смысл
    "цель": "🧭",
    "смысл": "🧭",
}

def format_prediction_inline(q: dict, display_name: str) -> str:
    quote = esc(strip_outer_quotes(q["quote"]))
    author = esc(q["author"].strip())

    source = esc((q.get("source") or "").strip())

    line2 = f"- {author}"
    if source:
        line2 += f', "{source}"'


    tag_norm = normalize_tag((q.get("tag") or "").strip())
    tag_line = emoji_for_tag(tag_norm) if tag_norm else ""

    header = f"🔮 Предсказание для {esc(display_name)}"

    # Telegram quote formatting:
    quote_block = f"<blockquote>{quote}</blockquote>"

    if tag_line:
        return f"{header}\n\n{quote_block}\n\n{line2} {esc(tag_line)}"
    return f"{header}\n\n{quote_block}\n\n{line2}"


def format_prediction(q: dict, message) -> str:
    username = message.from_user.username
    display_name = f"@{username}" if username else (message.from_user.first_name or "друг")

    date_str = datetime.now().strftime("%d.%m.%Y")
    header = f"🔮 Предсказание для {esc(display_name)}"

    quote = esc(strip_outer_quotes(q["quote"]))
    author = esc(q["author"].strip())

    source = esc((q.get("source") or "").strip())

    line2 = f"- {author}"
    if source:
     line2 += f', "{source}"'


    tag_norm = normalize_tag((q.get("tag") or "").strip())
    tag_line = emoji_for_tag(tag_norm) if tag_norm else ""

    quote_block = f"<blockquote>{quote}</blockquote>"

    if tag_line:
        return f"{header}\n\n{quote_block}\n\n{line2} {esc(tag_line)}"
    return f"{header}\n\n{quote_block}\n\n{line2}"

def format_chat_prediction(q: dict) -> str:
    quote = esc(strip_outer_quotes(q["quote"]))
    author = esc(q["author"].strip())
    source = esc((q.get("source") or "").strip())

    header = "🌞 Общее предсказание на сегодня для чата"

    line2 = f"- {author}"
    if source:
        line2 += f', "{source}"'

    tag_norm = normalize_tag((q.get("tag") or "").strip())
    tag_line = emoji_for_tag(tag_norm) if tag_norm else ""

    quote_block = f"<blockquote>{quote}</blockquote>"

    if tag_line:
        return f"{header}\n\n{quote_block}\n\n{line2} {esc(tag_line)}"
    return f"{header}\n\n{quote_block}\n\n{line2}"

def is_bot_mentioned(message) -> bool:
    # 1) Если Telegram прислал entities с упоминанием @username
    if message.entities:
        for ent in message.entities:
            if ent.type == "mention" and message.text:
                mention_text = message.text[ent.offset: ent.offset + ent.length]
                if mention_text.lower() == f"@{BOT_USERNAME.lower()}":
                    return True

    # 2) На всякий случай: простая проверка по тексту
    if message.text and f"@{BOT_USERNAME.lower()}" in message.text.lower():
        return True

    return False

def send_daily_chat_prediction(chat_id: int):
    if was_daily_sent_today(chat_id):
        return

    if not QUOTES:
        return

    q = pick_quote_bag(chat_id)
    text = format_chat_prediction(q)

    try:
        bot.send_message(chat_id, text, parse_mode="HTML")

        sticker_id = get_random_sticker_file_id()
        if sticker_id:
            bot.send_sticker(chat_id, sticker_id)

        mark_daily_sent(chat_id)
        

    except Exception as e:
        print(f"Failed to send daily prediction to chat {chat_id}: {e}")


def daily_scheduler():
    last_checked_date = None

    while True:
        try:
            now = datetime.now(BOT_TZ)
            today = now.strftime("%Y-%m-%d")
            hour = now.hour
            minute = now.minute

            # в 09:00 отправляем во все известные чаты
            if hour == 9 and minute == 0:
                # чтобы в рамках одной минуты не дублировать из-за цикла
                if last_checked_date != today:
                    for chat_id in list(known_chats):
                        send_daily_chat_prediction(chat_id)
                    last_checked_date = today

            time.sleep(30)

        except Exception as e:
            print(f"Scheduler error: {e}")
            time.sleep(30)

@bot.inline_handler(func=lambda query: True)
def inline_prediction(query):
    if not QUOTES:
        result = InlineQueryResultArticle(
            id=str(time.time()),
            title="База цитат пустая",
            description="Проверь quotes.xlsx",
            input_message_content=InputTextMessageContent("База цитат пустая. Проверь quotes.xlsx 🙂")
        )
        bot.answer_inline_query(query.id, [result], cache_time=1)
        return

    q = pick_quote_bag(query.from_user.id)

    # Важно: в инлайне у нас нет message, поэтому делаем текст без ника/даты,
    # либо заменяем на более общий заголовок.
    user = query.from_user
    name = f"@{user.username}" if user.username else user.first_name
    text = format_prediction_inline(q, name)


    tag_norm = normalize_tag(q.get('tag',''))
    desc = f"{emoji_for_tag(tag_norm)}" if tag_norm else "Тема: случайная"

    result = InlineQueryResultArticle(
        id=str(time.time()),
        title="Получить предсказание",
        description=desc,
        input_message_content=InputTextMessageContent(text, parse_mode="HTML")
    )

    bot.answer_inline_query(query.id, [result], cache_time=1)


@bot.message_handler(commands=["start", "help"])
def start_help(message):
    remember_chat(message.chat.id)

    bot.reply_to(
        message,
        "Привет! Отметь меня в сообщении, и я дам предсказание цитатой.\n"
        "Например: @{} дай предсказание".format(BOT_USERNAME)
    )

@bot.message_handler(func=lambda m: bool(m.text) and is_bot_mentioned(m))
def send_prediction_on_mention(message):
    if not QUOTES:
        bot.reply_to(message, "Похоже, база цитат пустая. Проверь файл quotes.xlsx 🙂")
        return

    remember_chat(message.chat.id)

    q = pick_quote_bag(message.chat.id)
    bot.reply_to(message, format_prediction(q, message), parse_mode="HTML")


if __name__ == "__main__":
    print("Bot is running...")

    scheduler_thread = threading.Thread(target=daily_scheduler, daemon=True)
    scheduler_thread.start()

    bot.infinity_polling(timeout=30, long_polling_timeout=30)